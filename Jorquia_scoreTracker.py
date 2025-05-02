import tkinter as tk
from openpyxl import Workbook, load_workbook
import os

filename = "student_scores.xlsx"

if not os.path.exists(filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Surname"
    sheet["B1"] = "Grade"
    sheet["C1"] = "Result"
    workbook.save(filename)

def create():
    surname = surname_entry.get()
    grade = grade_entry.get()

    if not surname or not grade:
        print("All fields are required.")
        return

    try:
        grades = int(grade)
    except ValueError:
        print("Grade must be a number.")
        return

    if 75 <= grades <= 100:
        result = "Passed"
    elif 0 <= grades < 75:
        result = "Failed"
    else:
        print("Grade must be between 0 and 100.")
        return

    workbook = load_workbook(filename)
    sheet = workbook.active
    sheet.append([surname, grades, result])
    workbook.save(filename)
    print("Data saved.")

root = tk.Tk()
root.title("Score tracker")
root.geometry("200x200")
root.configure(bg= "light blue")

surname_label = tk.Label(root, text="Surname")
surname_label.pack()

surname_entry = tk.Entry(root)
surname_entry.pack()

grade_label = tk.Label(root, text="Grade")
grade_label.pack()

grade_entry = tk.Entry(root)
grade_entry.pack()

check_button = tk.Button(root, text="Save Data", command=create)
check_button.pack(pady=10)

root.mainloop()
