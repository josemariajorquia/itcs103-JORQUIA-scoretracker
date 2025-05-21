import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl import Workbook

def validate_inputs():
    name = entry_name.get()
    course = entry_course.get()
    grade = entry_grade.get()

    if not name or not course or not grade:
        messagebox.showerror("Input Error", "All fields are required!")
        return False
    return True

def save_to_excel():
    if not validate_inputs():
        return

    name = entry_name.get()
    course = entry_course.get()
    grade = entry_grade.get()

    wb = load_workbook("Grade.xlsx")
    if "Grades" in wb.sheetnames:
        ws = wb["Grades"]
    else:
        ws = wb.create_sheet("Grades")
    ws.append([name, course, grade])
    wb.save("Grade.xlsx")

    messagebox.showinfo("Done", "Your entry has been recorded!")

    entry_name.delete(0, tk.END)
    entry_course.delete(0, tk.END)
    entry_grade.delete(0, tk.END)

def show_data():
    wb = load_workbook("Grade.xlsx")
    ws = wb["Grades"]

    data = tk.Toplevel(main_window)
    data.title("Student Data")
    data.geometry("250x180")
    data.configure(bg="light gray")

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        for j, value in enumerate(row):
            label = tk.Label(data, text=value, bg="light gray", padx=4, pady=2)
            label.grid(row=i, column=j)

main_window = tk.Tk()
main_window.geometry("300x300")
main_window.title("Grade Report")
main_window.configure(bg="light gray")

label_header = tk.Label(main_window, text="Grade Report", font=("arial", 14), bg="light gray")
label_header.pack(pady=15)

main_frame = tk.Frame(main_window, bg="light gray")
main_frame.pack()

form_frame = tk.Frame(main_frame, bg="light gray")
form_frame.pack(side="left", padx=10, pady=10)

label_name = tk.Label(form_frame, text="Name", font=("arial", 10), bg="light gray")
label_name.grid(row=0, column=0, sticky="w", pady=6)

entry_name = tk.Entry(form_frame, width=22)
entry_name.grid(row=0, column=1, pady=6)

label_course = tk.Label(form_frame, text="Course", font=("arial", 10), bg="light gray")
label_course.grid(row=1, column=0, sticky="w", pady=6)

entry_course = tk.Entry(form_frame, width=22)
entry_course.grid(row=1, column=1, pady=6)

label_grade = tk.Label(form_frame, text="Grade", font=("arial", 10), bg="light gray")
label_grade.grid(row=2, column=0, sticky="w", pady=6)

entry_grade = tk.Entry(form_frame, width=22)
entry_grade.grid(row=2, column=1, pady=6)

button_frame = tk.Frame(main_frame, bg="light gray")
button_frame.pack(side="left", padx=20, pady=10)

button_save = tk.Button(button_frame, text="Save", font=("arial", 9), width=12, command=save_to_excel)
button_save.pack(pady=10)

button_view = tk.Button(button_frame, text="View Data", font=("arial", 9), width=12, command=show_data)
button_view.pack(pady=10)

main_window.mainloop()